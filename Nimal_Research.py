import os
import re
import fitz
import mysql.connector
from customtkinter import CTkImage, CTkLabel
from PIL import Image
import pandas as pd
import tkinter as tk
from tkinter import ttk
import customtkinter as ctk
from tkinter import messagebox
from tkinter import filedialog
from openpyxl.reader.excel import load_workbook
from tkcalendar import DateEntry

conexao = mysql.connector.connect(
    host="192.168.0.101",
    user="seu_usuario",
    password="sua_senha",
    database="nimalnotas"
)

cursor = conexao.cursor()
def mostrar_visao_geral():
        global tree, frame2

        frame2.configure(width=1000, height=400)
        visao_frame = ctk.CTkFrame(frame2, fg_color="gray")
        visao_frame.place(relwidth=1, relheight=1)

        def carregar_dados(valor=None,coluna=None):
            for row in tree.get_children():
                tree.delete(row)

            # Conecta ao banco de dados
            conexao = mysql.connector.connect(
                host="192.168.0.101",
                user="seu_usuario",
                password="sua_senha",
                database="nimalnotas"
            )
            cursor = conexao.cursor()

            if conexao.is_connected():
                print("Conexão bem-sucedida ao MySQL")

            # SQL básico com filtro opcional
            sql_query = """
                SELECT local, orcamento,pedido,data, situacao, cliente, razao, representante, itens,total, vencimento, nf, valor_nf, distribuidor 
                FROM nimal
            """
            if valor:  # Se um valor for passado, filtra pela data
                sql_query += f" WHERE {coluna} LIKE %s "
                cursor.execute(sql_query, (valor,))
            else:  # Se não houver filtro, executa o SELECT sem condições
                cursor.execute(sql_query)

            # Carregar os resultados e exibir na Treeview
            resultados = cursor.fetchall()
            for i, linha in enumerate(resultados):
                tag = "odd" if i % 2 == 0 else "even"  # Alternando entre 'odd' e 'even'
                tree.insert("", tk.END, values=linha, tags=(tag,))

            # Fechar o cursor e a conexão
            conexao.commit()
            cursor.close()
            conexao.close()

        def aplicar_filtro():
            coluna_selecionada = combobox_colunas.get()  # Obtém o texto exibido na ComboBox
            coluna_real = opcoes.get(coluna_selecionada)  # Converte para o valor interno (nome real da coluna)

            valor_filtrado = entry_filtro.get().strip() + '%'

            if coluna_real:  # Verifica se a coluna foi selecionada corretamente
                # Ajusta a consulta para usar a coluna selecionada
                for row in tree.get_children():
                    tree.delete(row)

                conexao = mysql.connector.connect(
                    host="192.168.0.101",
                    user="seu_usuario",
                    password="sua_senha",
                    database="nimalnotas"
                )
                cursor = conexao.cursor()

                # Adicionando aspas ao redor dos nomes das colunas para evitar erro de sintaxe SQL
                sql_query = f"SELECT local, orcamento,pedido,data, situacao, cliente, razao, representante, itens,total, vencimento, nf, valor_nf, distribuidor FROM nimal WHERE `{coluna_real}` LIKE %s"
                cursor.execute(sql_query, (valor_filtrado,))
                resultados = cursor.fetchall()

                for i, linha in enumerate(resultados):
                    tag = "odd" if i % 2 == 0 else "even"
                    tree.insert("", tk.END, values=linha, tags=(tag,))

                cursor.close()
                conexao.close()
            else:
                messagebox.showerror("Erro", "Por favor, selecione uma coluna válida para filtrar.")

                # Estilo do cabeçalho do Treeview

        style = ttk.Style()
        style.theme_use("alt")

        # Configurações de estilo para o cabeçalho
        style.configure("Treeview.Heading",
                        font=("Arial",10),
                        background="#5946b4",
                        foreground="white",
                        padding=(10, 5))

        # Estilo do corpo do Treeview
        style.configure("Treeview",
                        fieldbackground="#EEEEEE",
                        foreground="black",
                        rowheight=50,
                        borderwidth=100,
                        relief="flat")

        # Criando a tabela com Treeview
        colunas = ("Local", "OGE", "Pedido", "Data", "Fase", "Cliente", "Razao", "Rep",
                   "Itens", "Total", "Venc", "NF", "Valor NF", "DTVM")

        # Criando o Treeview
        tree = ttk.Treeview(visao_frame, columns=colunas, show="headings")
        tree.pack(fill=tk.BOTH, expand=True)

        # Configurar as tags para as linhas
        tree.tag_configure("odd", background="#E6E6E6")  # Cor para linhas ímpares
        tree.tag_configure("even", background="#EEEEEE")  # Cor para linhas pares

        for coluna in colunas:
            tree.heading(coluna, text=coluna)
            tree.column(coluna, width=25)
            tree.column("Itens", width=2, anchor="center")
            tree.column("Cliente", width=18, anchor="center")
            tree.column("Fase", width=15, anchor="center")
            tree.column("Local", width=19, anchor="center")


        # Carregar dados ao inicializar
        carregar_dados()

        # Botão exportar
        def exportar_para_excel():
            caminho_arquivo = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
            if caminho_arquivo:
                # Extraindo os dados da Treeview para um DataFrame
                dados = [tree.item(item)["values"] for item in tree.get_children()]
                df = pd.DataFrame(dados, columns=colunas)

                # Usando ExcelWriter para definir a linha inicial de escrita
                with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, startcol=1)

                messagebox.showinfo("Exportação", f"Dados exportados com sucesso para {caminho_arquivo}")

        def duplicar_orcamento():
            try:
                global orcamento_selecionado

                # Obtendo o orcamento selecionado da Treeview
                orcamento_selecionado = tree.item(tree.selection())["values"][1]

                # Incrementar o valor do orçamento para registro
                base_orcamento = orcamento_selecionado.rsplit('-', 1)[0]
                numero_incremento = int(orcamento_selecionado.rsplit('-', 1)[1]) + 1
                novo_orcamento = f"{base_orcamento}-{numero_incremento}"

                cursor.execute(
                    "SELECT local, orcamento,pedido,data, situacao, cliente, razao, representante, itens,total, vencimento, nf, valor_nf, distribuidor FROM nimal WHERE orcamento = %s",
                    (orcamento_selecionado,))
                linha_selecionada = cursor.fetchone()

                print(linha_selecionada)

                # Copiar a linha selecionada e modificar o orçamento
                dados_copia = [
                    linha_selecionada[0],  # local
                    novo_orcamento,  # orcamento
                    linha_selecionada[2],  # pedido
                    linha_selecionada[3],  # data
                    linha_selecionada[4],  # situacao
                    linha_selecionada[5],  # cliente
                    linha_selecionada[6],  # razao
                    linha_selecionada[7],  # representante
                    linha_selecionada[8],  # itens
                    linha_selecionada[9],  # total
                    linha_selecionada[10],  # vencimento
                    linha_selecionada[11],  # nf
                    linha_selecionada[12],  # valor_nf
                    linha_selecionada[13]  # distribuidor
                ]

                print(dados_copia)

                # Inserir a nova linha na tabela no banco de dados
                cursor.execute(
                    "INSERT INTO nimal (local,orcamento, pedido, data, situacao, cliente, razao, representante, itens, total, vencimento, nf, valor_nf, distribuidor) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                    dados_copia
                )
                conexao.commit()
                carregar_dados()
                atualizar_contagem()

            except Exception as e:
                messagebox.showerror("Erro", "Por favor,selecione um pedido valido para duplicar.Certifique-se que o pedido selecionado não apresenta uma duplicata.")

        def editar_dados():
            # Verifica se há uma linha selecionada
            item_selecionado = tree.selection()
            if not item_selecionado:
                messagebox.showwarning("Aviso", "Por favor, selecione uma linha para editar.")
                return

            # Recupera os valores da linha selecionada
            valores_atuais = tree.item(item_selecionado, "values")

            janela_edicao = ctk.CTkToplevel()
            janela_edicao.title("Teste")
            janela_edicao.geometry("600x800")
            janela_edicao.resizable(width=False, height=False)


            # frames
            campos = [
                "local","orcamento","pedido", "data", "situacao", "cliente", "razao",
            "representante", "itens", "total", "vencimento", "nf", "valor_nf", "distribuidor"
            ]

            entradas = {}

            frameA = ctk.CTkFrame(janela_edicao, fg_color="#EEEEEE", width=700, height=700, corner_radius=10)
            frameA.place(relx=0.5, rely=0.43, anchor='center')

            frameC = ctk.CTkFrame(frameA, width=700, height=300, corner_radius=10, fg_color="#5946b4")
            frameC.place(relx=0.5, rely=0.1, anchor='center')

            frameB = ctk.CTkFrame(frameA, fg_color="#EEEEEE", width=600, height=900, corner_radius=10)
            frameB.place(relx=0.5, rely=0.8, anchor='center')

            # labels e entries

            logo_img_data = Image.open("nimall2.png")
            logo_img = CTkImage(dark_image=logo_img_data, light_image=logo_img_data, size=(90, 90))
            img = ctk.CTkLabel(master=frameC, text="", image=logo_img)
            img.place(relx=0.8, rely=0.45, anchor='center')

            titulo = ctk.CTkLabel(frameC, text="Editar Pedido", font=("Arial Black", 24), text_color="white")
            titulo.place(relx=0.1, rely=0.4, anchor='w')

            idlabel = ctk.CTkLabel(frameC, text=f"OGE: {valores_atuais[1]}", font=("Arial Black", 18), text_color="white")
            idlabel.place(relx=0.1, rely=0.5, anchor='w')


            CTkLabel(frameA, text="Local", font=("Arial", 12), text_color="black").place(relx=0.12, rely=0.21,
                                                                                            anchor='w')
            CTkLabel(frameA, text="Pedido", font=("Arial", 12), text_color="black").place(relx=0.52, rely=0.21,
                                                                                          anchor='w')
            CTkLabel(frameA, text="Data", font=("Arial", 12), text_color="black").place(relx=0.12, rely=0.31,
                                                                                           anchor='w')
            CTkLabel(frameA, text="Fase", font=("Arial", 12), text_color="black").place(relx=0.52, rely=0.31,
                                                                                                  anchor='w')
            CTkLabel(frameA, text="Cliente", font=("Arial", 12), text_color="black").place(relx=0.12, rely=0.41,
                                                                                               anchor='w')
            CTkLabel(frameA, text="Razao", font=("Arial", 12), text_color="black").place(relx=0.52, rely=0.41,
                                                                                         anchor='w')
            CTkLabel(frameA, text="Representante", font=("Arial", 12), text_color="black").place(relx=0.12, rely=0.51,
                                                                                     anchor='w')
            CTkLabel(frameA, text="Itens", font=("Arial", 12), text_color="black").place(relx=0.52, rely=0.51,
                                                                                                 anchor='w')
            CTkLabel(frameA, text="Total", font=("Arial", 12), text_color="black").place(relx=0.12, rely=0.61,
                                                                                                 anchor='w')
            CTkLabel(frameA, text="Vencimento", font=("Arial", 12), text_color="black").place(relx=0.52, rely=0.61,
                                                                                                 anchor='w')
            CTkLabel(frameA, text="Nota Fiscal", font=("Arial", 12), text_color="black").place(relx=0.12, rely=0.71,
                                                                                                 anchor='w')
            CTkLabel(frameA, text="Valor nota Fiscal", font=("Arial", 12), text_color="black").place(relx=0.52, rely=0.71,
                                                                                                 anchor='w')
            CTkLabel(frameA, text="Distribuidor", font=("Arial", 12), text_color="black").place(relx=0.12, rely=0.81,
                                                                                                 anchor='w')

            # Entrys
            # Entrys baseadas na lista "campos"
            local_opcoes = ["NIMAL TECNOLOGIA", "NIMAL WORLD"]
            local_entry = ctk.CTkComboBox(frameA, values=local_opcoes, button_color="#5946b4", border_color="#5946b4",
                                          font=("Arial", 12), width=250, height=30)
            local_entry.place(relx=0.3, rely=0.25, anchor='center')
            local_entry.set(valores_atuais[0])  # Valor inicial do campo
            entradas["local"] = local_entry

            orcamento_entry = ctk.CTkEntry(frameA,font=("Arial", 12), width=0, border_color="#5946b4")
            orcamento_entry.place(relx=17.3, rely=17.1, anchor='center')
            orcamento_entry.insert(0, valores_atuais[1])
            entradas["orcamento"] = orcamento_entry


            pedido_entry = ctk.CTkEntry(frameA, font=("Arial", 12), width=250, border_color="#5946b4")
            pedido_entry.place(relx=0.7, rely=0.25, anchor='center')
            pedido_entry.insert(0, valores_atuais[2])  # Valor inicial do campo
            entradas["pedido"] = pedido_entry

            data_entry = DateEntry(frameA, width=18, font=("Arial", 12), date_pattern="dd/mm/yyyy")
            data_entry.place(relx=0.25, rely=0.35, anchor='center')
            data_entry.set_date(valores_atuais[3])  # Valor inicial do campo
            entradas["data"] = data_entry

            fase_opcoes = ["(Aberto)", "Faturado", "Cancelado"]
            fase_entry = ctk.CTkComboBox(frameA, values=fase_opcoes, button_color="#5946b4", border_color="#5946b4",
                                         font=("Arial", 12), width=250, height=30)
            fase_entry.place(relx=0.7, rely=0.35, anchor='center')
            fase_entry.set(valores_atuais[4])  # Valor inicial do campo
            entradas["situacao"] = fase_entry

            cliente_entry = ctk.CTkEntry(frameA, font=("Arial", 12), width=250, border_color="#5946b4")
            cliente_entry.place(relx=0.3, rely=0.45, anchor='center')
            cliente_entry.insert(0, valores_atuais[5])  # Valor inicial do campo
            entradas["cliente"] = cliente_entry

            razao_entry = ctk.CTkEntry(frameA, font=("Arial", 12), width=250, border_color="#5946b4")
            razao_entry.place(relx=0.7, rely=0.45, anchor='center')
            razao_entry.insert(0, valores_atuais[6])  # Valor inicial do campo
            entradas["razao"] = razao_entry

            representante_entry = ctk.CTkEntry(frameA, font=("Arial", 12), width=250, border_color="#5946b4")
            representante_entry.place(relx=0.3, rely=0.55, anchor='center')
            representante_entry.insert(0, valores_atuais[7])  # Valor inicial do campo
            entradas["representante"] = representante_entry

            itens_entry = ctk.CTkEntry(frameA, font=("Arial", 12), width=250, border_color="#5946b4")
            itens_entry.place(relx=0.7, rely=0.55, anchor='center')
            itens_entry.insert(0, valores_atuais[8])  # Valor inicial do campo
            entradas["itens"] = itens_entry

            total_entry = ctk.CTkEntry(frameA, font=("Arial", 12), width=250, border_color="#5946b4")
            total_entry.place(relx=0.3, rely=0.65, anchor='center')
            total_entry.insert(0, valores_atuais[9])  # Valor inicial do campo
            entradas["total"] = total_entry

            vencimento_entry = DateEntry(frameA, width=18, font=("Arial", 12), date_pattern="dd/mm/yyyy")
            vencimento_entry.place(relx=0.65, rely=0.65, anchor='center')
            entradas["vencimento"] = vencimento_entry

            nf_entry = ctk.CTkEntry(frameA, font=("Arial", 12), width=250, border_color="#5946b4")
            nf_entry.place(relx=0.3, rely=0.75, anchor='center')
            nf_entry.insert(0, valores_atuais[11])  # Valor inicial do campo
            entradas["nf"] = nf_entry

            valor_nf_entry = ctk.CTkEntry(frameA, font=("Arial", 12), width=250, border_color="#5946b4")
            valor_nf_entry.place(relx=0.7, rely=0.75, anchor='center')
            valor_nf_entry.insert(0, valores_atuais[12])  # Valor inicial do campo
            entradas["valor_nf"] = valor_nf_entry

            distr_entry = ctk.CTkEntry(frameA, font=("Arial", 12), width=250, border_color="#5946b4")
            distr_entry.place(relx=0.3, rely=0.85, anchor='center')
            distr_entry.insert(0, valores_atuais[13])  # Valor inicial do campo
            entradas["distribuidor"] = distr_entry



            # Função para salvar as edições
            def confirmar_edicoes():
                novos_valores = [entradas[campo].get() for campo in campos]

                # Atualizar a Treeview
                tree.item(item_selecionado, values=novos_valores)

                # Atualizar o banco de dados
                conexao = mysql.connector.connect(
                    host="192.168.0.101",
                    user="seu_usuario",
                    password="sua_senha",
                    database="nimalnotas"
                )
                cursor = conexao.cursor()

                try:
                    sql_update = """
                        UPDATE nimal
                        SET local = %s, orcamento = %s, pedido = %s, data = %s, situacao = %s, cliente = %s, 
                            razao = %s, representante = %s, itens = %s, total = %s, vencimento = %s, 
                            nf = %s, valor_nf = %s, distribuidor = %s
                        WHERE orcamento = %s
                    """
                    # Substitui os parâmetros na consulta SQL
                    cursor.execute(sql_update, (
                    *novos_valores, valores_atuais[1]))  # Usar o valor antigo de "orcamento" na cláusula WHERE
                    conexao.commit()
                    messagebox.showinfo("Sucesso", "Dados atualizados com sucesso!")
                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao atualizar dados: {e}")
                finally:
                    cursor.close()
                    conexao.close()

                atualizar_contagem()
                janela_edicao.destroy()

            # Botão para confirmar as alterações
            ctk.CTkButton(
                janela_edicao, text="Confirmar", command=confirmar_edicoes,
                fg_color="#5946b4", hover_color="#4b3b96", width=200, height=40, font=("Arial", 14)
            ).place(relx=0.3, rely=0.95, anchor=tk.CENTER)

            # Botão para cancelar a edição
            ctk.CTkButton(
                janela_edicao, text="Cancelar", command=janela_edicao.destroy,
                fg_color="#5946b4", hover_color="#4b3b96", width=200, height=40, font=("Arial", 14)
            ).place(relx=0.7, rely=0.95, anchor=tk.CENTER)

        def remover_dados():
            global orcamento_selecionado

            orcamento_selecionado = tree.item(tree.selection())["values"][1]

            resposta = messagebox.askyesno("Confirmação",
                                           f"Tem certeza que deseja remover o pedido {orcamento_selecionado} do banco de dados?")
            if resposta:
                try:
                    conexao = mysql.connector.connect(
                        host="192.168.0.101",
                        user="seu_usuario",
                        password="sua_senha",
                        database="nimalnotas"
                    )
                    cursor = conexao.cursor()

                    # Remover a linha selecionada
                    sql_delete = "DELETE FROM nimal WHERE orcamento = %s"
                    cursor.execute(sql_delete, (orcamento_selecionado,))

                    conexao.commit()
                    cursor.close()
                    conexao.close()
                    carregar_dados()
                    atualizar_contagem()


                except Exception as e:
                    messagebox.showerror("Erro", f"Erro ao remover dados: {e}")

        def contar_elementos():
            # Conexão com o banco de dados
            conexao = mysql.connector.connect(
                host="192.168.0.101",
                user="seu_usuario",
                password="sua_senha",
                database="nimalnotas"
            )
            cursor = conexao.cursor()

            # Executando o comando SQL para contar os elementos
            query = "SELECT COUNT(*) FROM nimal"  # Substitua 'projetos' pelo nome da sua tabela
            cursor.execute(query)
            resultado = cursor.fetchone()[0]

            # Fechando a conexão
            cursor.close()
            conexao.close()


            return resultado

        def contar_concluidos():
            # Conexão com o banco de dados
            conexao = mysql.connector.connect(
                host="192.168.0.101",
                user="seu_usuario",
                password="sua_senha",
                database="nimalnotas"
            )
            cursor = conexao.cursor()

            # Executando o comando SQL para contar os elementos com "Status = Concluído"
            query = "SELECT COUNT(*) FROM nimal WHERE situacao = 'Faturado'"  # Ajuste 'projetos' e 'status' se necessário
            cursor.execute(query)
            resultado = cursor.fetchone()[0]

            # Fechando a conexão
            cursor.close()
            conexao.close()


            return resultado

        def contar_em_aberto():
            # Conexão com o banco de dados
            conexao = mysql.connector.connect(
                host="192.168.0.101",
                user="seu_usuario",
                password="sua_senha",
                database="nimalnotas"
            )
            cursor = conexao.cursor()

            # Executando o comando SQL para contar os elementos com "Status = Em Aberto"
            query = "SELECT COUNT(*) FROM nimal WHERE vencimento IS NULL"  # Ajuste 'projetos' e 'status' se necessário
            cursor.execute(query)
            resultado = cursor.fetchone()[0]

            # Fechando a conexão
            cursor.close()
            conexao.close()


            return resultado

        def atualizar_contagem():
            try:
                conexao = mysql.connector.connect(
                    host="192.168.0.101",
                    user="seu_usuario",
                    password="sua_senha",
                    database="nimalnotas"
                )
                cursor = conexao.cursor()

                cursor.execute("SELECT COUNT(*) FROM projetos WHERE status = 'Concluído'")
                total_concluidos = cursor.fetchone()[0]

                cursor.execute("SELECT COUNT(*) FROM projetos WHERE status = 'Em Aberto'")
                total_em_aberto = cursor.fetchone()[0]

                cursor.execute("SELECT COUNT(*) FROM projetos;")
                total_elementos = cursor.fetchone()[0]


                label1.configure(text=f"Total: {total_elementos}")
                label2.configure(text=f"Concluídos: {total_concluidos}")
                label3.configure(text=f"Em Aberto: {total_em_aberto}")

            except mysql.connector.Error as e:
                print(f"Erro ao atualizar contagem: {e}")
            finally:
                if cursor:
                    cursor.close()
                if conexao:
                    conexao.close()

        def alterar_dados(venc, nf, dist, valor, fase):
            global orcamento_selecionado

            orcamento_selecionado = tree.item(tree.selection())["values"][1]

            janela_edicao = ctk.CTkToplevel()
            janela_edicao.title("Adcionar Dados")
            janela_edicao.geometry("600x500")
            janela_edicao.resizable(width=False, height=False)

            frameA = ctk.CTkFrame(janela_edicao, fg_color="#EEEEEE", width=700, height=700, corner_radius=10)
            frameA.place(relx=0.5, rely=0.7, anchor='center')

            frameC = ctk.CTkFrame(frameA, width=700, height=300, corner_radius=10, fg_color="#5946b4")
            frameC.place(relx=0.5, rely=0.1, anchor='center')

            frameB = ctk.CTkFrame(frameA, fg_color="#EEEEEE", width=600, height=900, corner_radius=10)
            frameB.place(relx=0.5, rely=0.8, anchor='center')

            logo_img_data = Image.open("nimall2.png")
            logo_img = CTkImage(dark_image=logo_img_data, light_image=logo_img_data, size=(90, 90))
            img = ctk.CTkLabel(master=frameC, text="", image=logo_img)
            img.place(relx=0.8, rely=0.45, anchor='center')

            titulo = ctk.CTkLabel(frameC, text="Editar Pedido", font=("Arial Black", 24), text_color="white")
            titulo.place(relx=0.1, rely=0.4, anchor='w')

            idlabel = ctk.CTkLabel(frameC, text="Adicionar", font=("Arial Black", 18),
                                   text_color="white")
            idlabel.place(relx=0.1, rely=0.5, anchor='w')

            # Campos para edição
            ctk.CTkLabel(frameA, text="Vencimento:",).place(relx=0.12, rely=0.21,anchor='w')
            entrada_venc = DateEntry(frameA, width=26, font=("Arial", 12), date_pattern="dd/mm/yyyy")
            entrada_venc.insert(0, venc)
            entrada_venc.place(relx=0.3, rely=0.25, anchor='center')

            ctk.CTkLabel(frameA, text="Nº da NF:").place(relx=0.52, rely=0.21,anchor='w')
            entrada_nf = ctk.CTkEntry(frameA, font=("Arial", 12), width=250, border_color="#5946b4")
            entrada_nf.insert(0, nf)
            entrada_nf.place(relx=0.7, rely=0.25, anchor='center')

            ctk.CTkLabel(frameA, text="Distribuidor:").place(relx=0.12, rely=0.31,anchor='w')
            entrada_dist = ctk.CTkEntry(frameA, font=("Arial", 12), width=250, border_color="#5946b4")
            entrada_dist.insert(0, dist)
            entrada_dist.place(relx=0.3, rely=0.35, anchor='center')

            ctk.CTkLabel(frameA, text="Valor:").place(relx=0.52, rely=0.31,anchor='w')
            entrada_valor = ctk.CTkEntry(frameA, font=("Arial", 12), width=250, border_color="#5946b4")
            entrada_valor.insert(0, valor)
            entrada_valor.place(relx=0.7, rely=0.35, anchor='center')

            ctk.CTkLabel(frameA, text="Fase:").place(relx=0.12, rely=0.41,anchor='w')
            fase_opcoes = ["(Aberto)", "Faturado", "Cancelado"]
            entrada_fase = ctk.CTkComboBox(frameA, values=fase_opcoes, button_color="#5946b4", border_color="#5946b4",
                                         font=("Arial", 12), width=250, height=30)
            entrada_fase.place(relx=0.3, rely=0.45, anchor='center')

            # Botão de confirmação para salvar as alterações
            def confirmar_edicao():
                # Conectar ao banco de dados
                conexao = mysql.connector.connect(
                    host="192.168.0.101",
                    user="seu_usuario",
                    password="sua_senha",
                    database="nimalnotas"
                )
                cursor = conexao.cursor()

                # Atualizar os dados no banco de dados
                sql_update = """
                    UPDATE nimal
                    SET vencimento = %s, nf = %s, distribuidor = %s, valor_nf = %s , situacao = %s
                    WHERE orcamento = %s
                """
                cursor.execute(sql_update, (
                    entrada_venc.get(), entrada_nf.get(), entrada_dist.get(), entrada_valor.get(), entrada_fase.get(),
                    orcamento_selecionado))

                conexao.commit()
                cursor.close()
                conexao.close()
                carregar_dados()
                atualizar_contagem()
                # Fechar a janela de edição
                janela_edicao.destroy()

                # Recarregar a visão geral

            ctk.CTkButton(
                janela_edicao, text="Confirmar", command=confirmar_edicao,
                fg_color="#5946b4", hover_color="#4b3b96", width=200, height=40, font=("Arial", 14)
            ).place(relx=0.5, rely=0.8, anchor=tk.CENTER)

        def extrair_informacoes_pdf(caminho_pdf):
            try:
                with fitz.open(caminho_pdf) as pdf:
                    conteudo = ""
                    for pagina in pdf:
                        conteudo += pagina.get_text()

                    linhas = conteudo.splitlines()

                    # Listas de padrões para buscar diferentes variações
                    padroes_venc = [r"Venc\.\s*(.*)", r"VENCIMENTO\s*(.*)"]
                    padroes_nf = [r"Nº\.\s*(.*)", r"Nº\s*(.*)", r"Número Fiscal\s*(.*)"]
                    padroes_dist = [r"IDENTIFICAÇÃO DO EMITENTE\s*(.*)", r"Distribuidor\s*(.*)", r"RECEBEMOS DE\s*(.*)"]
                    padroes_valor = [r"V\. TOTAL DA NOTA\s*(.*)", r"VALOR TOTAL DA NOTA\s*(.*)", r"V\. Total\s*(.*)"]
                    padroes_fase = ""

                    # Função auxiliar para tentar vários padrões até encontrar uma correspondência
                    def buscar_texto(padroes):
                        for padrao in padroes:
                            match = re.search(padrao, conteudo)
                            if match:
                                return match.group(1).strip()
                        return ""

                    # Usar a função auxiliar para buscar cada campo
                    venc_extraido = buscar_texto(padroes_venc)
                    nf_extraido = buscar_texto(padroes_nf)
                    dist_extraido = buscar_texto(padroes_dist)
                    valor_extraido = buscar_texto(padroes_valor)
                    fase_extraido = buscar_texto(padroes_fase)

                    # Chamar função para editar e confirmar dados
                    alterar_dados(venc_extraido, nf_extraido, dist_extraido, valor_extraido, fase_extraido)

            except Exception as e:
                messagebox.showerror("Erro", f"Erro ao processar o PDF: {e}")

        def selecionar_pdf():
            item_selecionado = tree.selection()
            if not item_selecionado:
                messagebox.showwarning("Aviso", "Por favor, selecione um pedido para adicionar a nota fiscal.")
                return

            caminho_pdf = filedialog.askopenfilename(filetypes=[("Arquivos PDF", "*.pdf")])
            if caminho_pdf:
                extrair_informacoes_pdf(caminho_pdf)

        logo_img_data = Image.open("nimall2.png")
        logo_img = CTkImage(dark_image=logo_img_data, light_image=logo_img_data, size=(130,130))
        CTkLabel(master=frame1, text="", image=logo_img).pack(pady=(100, 120), anchor="center")

        img1 = Image.open("Product_Documents.png")
        img1 = CTkImage(dark_image=img1, light_image=img1, size=(45,45))
        CTkLabel(master=frameA, text="", image=img1).place(x=20,rely=0.2)

        img1 = Image.open("Check Dollar.png")
        img1 = CTkImage(dark_image=img1, light_image=img1, size=(45, 45))
        CTkLabel(master=frameB, text="", image=img1).place(x=20, rely=0.2)

        img1 = Image.open("Cancel Subscription.png")
        img1 = CTkImage(dark_image=img1, light_image=img1, size=(45, 45))
        CTkLabel(master=frameC, text="", image=img1).place(x=20, rely=0.2)

        total_elementos = contar_elementos()
        total_concluidos = contar_concluidos()
        total_em_aberto = contar_em_aberto()

        # Exibindo o resultado
        label1 = tk.Label(frameA, text=f"Total: {total_elementos}",
                         font=("Arial Black", 14), fg="white",bg="#5946b4")
        label1.place(x=80, rely=0.25)

        label2 = tk.Label(frameB, text=f"Faturados: {total_concluidos}",
                         font=("Arial Black", 14), fg="white", bg="#5946b4")
        label2.place(x=80, rely=0.25)

        label3 = tk.Label(frameC, text=f"Sem Nota: {total_em_aberto}",
                          font=("Arial Black", 14), fg="white", bg="#5946b4")
        label3.place(x=80, rely=0.25)

        entry_filtro = ctk.CTkEntry(frameD, width=400, font=("Arial", 12),border_color="#5946b4")
        entry_filtro.place(x=40, rely=0.3)

        #botões

        botao_editar = ctk.CTkButton(frame1, text="Editar", command=editar_dados, fg_color="#5946b4", hover_color="#4b3b96",
                                     width=200, height=50, font=("Arial Bold", 16),image=imgb1,anchor="w")
        botao_editar.pack(pady=5, padx=5)


        botao_remover = ctk.CTkButton(frame1, text="Remover", command=remover_dados, fg_color="#5946b4",
                                      hover_color="#4b3b96", width=200, height=50, font=("Arial Bold", 16), image=imgb2,anchor="w")
        botao_remover.pack(pady=5, padx=5)


        botao_exportar = ctk.CTkButton(frame1, text="Gerar Relatório", command=exportar_para_excel, fg_color="#5946b4",
                                       hover_color="#4b3b96", width=200, height=50, font=("Arial Bold", 16),image=imgb4,anchor="w")
        botao_exportar.pack(pady=5, padx=(20,20))

        botao_duplicar = ctk.CTkButton(frame1, text="Duplicar", command=duplicar_orcamento, fg_color="#5946b4",
                                       hover_color="#4b3b96", width=200, height=50, font=("Arial Bold", 16),
                                       image=imgb8, anchor="w")
        botao_duplicar.pack(pady=5, padx=(20, 20))

        botao_extrair = ctk.CTkButton(frame1, text="Selecionar Nota", command=selecionar_pdf, fg_color="#5946b4",
                                       hover_color="#4b3b96", width=200, height=50, font=("Arial Bold", 16),
                                       image=imgb6, anchor="w")
        botao_extrair.pack(pady=5, padx=(20, 20))

        botao_importar = ctk.CTkButton(frame1, text="Importar Dados", command=importar_dados_excel, fg_color="#5946b4",
                                      hover_color="#4b3b96", width=200, height=50, font=("Arial Bold", 16),
                                      image=imgb7, anchor="w")
        botao_importar.pack(pady=5, padx=(20, 20))


        botao_filtrar = ctk.CTkButton(frameD, text="Pesquisar", command=aplicar_filtro, fg_color="#5946b4",
                                      hover_color="#4b3b96", width=200, height=30, font=("Arial Bold", 16),image=imgb5,anchor="w")
        botao_filtrar.place(x=800, rely=0.5, anchor="center")


        opcoes = {
            "Local": "local",
            "Orçamento": "orcamento",
            "Pedido": "pedido",
            "Data": "data",
            "Fase": "fase",
            "Cliente": "cliente",
            "Razão": "razao",
            "Horas": "horas",
            "Representante": "representante",
            "Itens": "itens",
            "Total": "total",
            "Vencimento": "vencimento",
            "Nota Fiscal": "nf",
            "Valor da Nota": "valor_nf",
            "Distribuidor": "distribuidor"
        }


        combobox_colunas = ctk.CTkComboBox(
            frameD,
            values=list(opcoes.keys()),  # Exibe as chaves do dicionário
            fg_color="#EEEEEE",
            border_color="#5946b4",
            width=200,
            text_color="black",
            button_color="#5946b4",
            button_hover_color="#4b3b96",
            dropdown_hover_color="#EEEEEE"
        )
        combobox_colunas.place(x=470, rely=0.3)
        combobox_colunas.set("Selecione uma opção")
def importar_dados_excel():
    caminho_excel = filedialog.askopenfilename(title="Selecione o arquivo Excel", filetypes=[("Excel files", "*.xlsx")])
    if not caminho_excel or not os.access(caminho_excel, os.R_OK):
        messagebox.showerror("Erro", "Arquivo inválido ou sem permissão de leitura.")
        return

    try:
        wb = load_workbook(caminho_excel)
        sheet = wb.active

        # Verificar e adicionar colunas necessárias
        if 'vencimento' not in sheet.cell(row=1, column=sheet.max_column).value:
            sheet.cell(row=1, column=sheet.max_column + 1, value="vencimento")
        if 'nf' not in sheet.cell(row=1, column=sheet.max_column).value:
            sheet.cell(row=1, column=sheet.max_column + 1, value="nf")
        if 'valor_nf' not in sheet.cell(row=1, column=sheet.max_column).value:
            sheet.cell(row=1, column=sheet.max_column + 1, value="valor_nf")
        if 'distribuidor' not in sheet.cell(row=1, column=sheet.max_column).value:
            sheet.cell(row=1, column=sheet.max_column + 1, value="distribuidor")

        # Conectar ao banco de dados
        conexao = mysql.connector.connect(
            host="192.168.0.101",
            user="seu_usuario",
            password="sua_senha",
            database="nimalnotas"
        )
        cursor = conexao.cursor()

        # Percorrer todas as linhas de dados do Excel (começando pela segunda linha)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Verificar se a linha possui colunas suficientes
            if len(row) < 14:
                continue

            # Atribuir valores com segurança
            local = row[1]
            orcamento = row[2]
            pedido = row[3]
            data = row[4]
            situacao = row[5]
            cliente = row[6]
            razao = row[7]
            representante = row[8]
            itens = row[9]
            total = row[10]
            vencimento = row[11]
            valor_nf = row[12]
            nf = row[13]
            distribuidor = row[14]

            # Verificar se o orçamento já existe no banco de dados
            cursor.execute("SELECT COUNT(*) FROM nimal WHERE orcamento = %s ", (orcamento,))
            resultado = cursor.fetchone()

            if resultado and resultado[0] > 0:
                # Atualizar se o orçamento já existe
                sql_update = """
                    UPDATE nimal
                    SET local = %s, pedido = %s, situacao = %s, cliente = %s, razao = %s,
                        representante = %s, itens = %s, total = %s, vencimento = %s,
                        valor_nf = %s, nf = %s, distribuidor = %s, data = %s
                    WHERE orcamento = %s
                """
                cursor.execute(sql_update, (
                    local, pedido, situacao, cliente, razao, representante, itens, total, vencimento, valor_nf, nf,
                    distribuidor, data, orcamento))
            else:
                # Inserir novo registro
                sql_insert = """
                    INSERT INTO nimal (
                        local, orcamento, pedido, situacao, cliente, razao, representante,
                        itens, total, vencimento, nf, valor_nf, distribuidor, data
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                cursor.execute(sql_insert, (
                    local, orcamento, pedido, situacao, cliente, razao, representante, itens, total, vencimento, nf,
                    valor_nf, distribuidor, data))

        conexao.commit()
        cursor.close()
        conexao.close()
        messagebox.showinfo("Sucesso", "Dados do Excel importados com sucesso.")

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao importar dados do Excel: {e}")
def get_screen_size():
    screen_width = janela.winfo_screenwidth()
    screen_height = janela.winfo_screenheight()
    return screen_width, screen_height
def centralizar_janela(largura, altura):
    # Obter dimensões da tela
    largura_tela = janela.winfo_screenwidth()
    altura_tela = janela.winfo_screenheight()

    # Calcular posição x e y para centralizar
    pos_x = (largura_tela - largura) // 2
    pos_y = (altura_tela - altura) // 2

    return f"{largura}x{altura}+{pos_x}+{pos_y}"


janela = ctk.CTk()
screen_width, screen_height = get_screen_size()
dimensoes = centralizar_janela(screen_width, screen_height)
janela.geometry(dimensoes)
janela.resizable(True, True)
janela.title("NimalResearch")
janela.state('zoomed')
janela.iconbitmap("nimal.ico")
ctk.set_appearance_mode("light")

imagem_pdf = ctk.CTkImage(Image.open("pdf.png"), size=(30, 30))
imagem_logo = ctk.CTkImage(Image.open("nimal.ico"), size=(80, 80))
imgb1 = ctk.CTkImage(Image.open("icons8-editar-50.png"), size=(30, 30))
imgb2 = ctk.CTkImage(Image.open("icons8-apagar-para-sempre-24.png"), size=(30, 30))
imgb3 = ctk.CTkImage(Image.open("icons8-adicionar-50.png"), size=(30, 30))
imgb4 = ctk.CTkImage(Image.open("icons8-lista-de-arquivo-de-peças-30.png"), size=(30, 30))
imgb5 = ctk.CTkImage(Image.open("icons8-pesquisar-64.png"), size=(20, 20))
imgb6 = ctk.CTkImage(Image.open("icons8-duplicata-50.png"), size=(30, 30))
imgb7 = ctk.CTkImage(Image.open("icons8-importar-30.png"), size=(30, 30))
imgb8 = ctk.CTkImage(Image.open("icons8-adicionar-arquivo-32.png"), size=(30, 30))

frame1 = ctk.CTkFrame(janela, fg_color="#5946b4", width=100, height=650,corner_radius=0)
frame1.pack(fill="y", anchor="w", side="left")

frame3 = ctk.CTkFrame(janela,fg_color="#E6E6E6", width=1000, height=220, corner_radius=10)
frame3.place(relx=0.6, rely=0.18, anchor='center')

frameA = ctk.CTkFrame(frame3,fg_color="#5946b4", width=300, height=70, corner_radius=10)
frameA.place(relx=0.18, rely=0.4, anchor='center')

frameB = ctk.CTkFrame(frame3,fg_color="#5946b4", width=300, height=70, corner_radius=10)
frameB.place(relx=0.50, rely=0.4, anchor='center')

frameC = ctk.CTkFrame(frame3,fg_color="#5946b4", width=300, height=70, corner_radius=10)
frameC.place(relx=0.82, rely=0.4, anchor='center')

frameD = ctk.CTkFrame(frame3,fg_color="#EEEEEE", width=950, height=70, corner_radius=10)
frameD.place(relx=0.5, rely=0.8, anchor='center')

frame2 = ctk.CTkFrame(janela, fg_color="#33415c", width=500, height=280, corner_radius=10)
frame2.place(relx=0.6, rely=0.68, anchor='center')

(CTkLabel(master=frame3, text="Gestão de Pedidos", font=("Arial Black", 30), text_color="#5946b4")
.place(x = 30 , y = 10))

mostrar_visao_geral()

janela.mainloop()
